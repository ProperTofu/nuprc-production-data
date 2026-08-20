#!/usr/bin/env python3
"""Nigeria's monthly rig count, from the Baker Hughes worldwide workbook.

NUPRC publishes rig disposition too, but quarterly and late: as of August 2026
the newest covered January to March, released on 10 April. Baker Hughes
publishes monthly, splits land from offshore, and ships an Excel file rather
than a watermarked PDF, so it is the better series for anything asking what
rigs are doing now.

Baker Hughes provide the rig count as a public service and ask to be credited
for it. Anything built on this file should say where the numbers came from
when it states its sources.

Discovery is by filename, not by URL. The workbook sits behind an opaque
/static-files/<uuid> link whose uuid changes when a new month is published,
so the page is scanned, each candidate's Content-Disposition is read, and the
one calling itself a WorldWide Rig Count Report is taken. Bookmarking the URL
would silently keep serving July forever.

Two workbooks are read, not one. The current report only reaches back to
January 2024; everything before that is in the separate "International Rig
Counts" file, which stores each month's WEEKLY counts summed rather than the
monthly average -- January 2024 land reads 37 there against 9 in the current
report, and 37 over 4 Fridays is 9.25. Dividing by the FridayCount column
reproduces the published figure for every overlapping month. Where they
overlap the current report wins, being published rather than derived.

Output is one continuous file rather than one per year, since the series is
continuous and a year boundary would be invented here rather than found in
the data. Inland waters is counted alongside land and offshore: Nigeria ran
one inland rig a month through most of 2024 and more in earlier years, so a
total of land plus offshore alone understates it.

Needs openpyxl. Run it through scripts/scrape_rig_count.sh, which installs
that into a throwaway container.

Usage:
    python scripts/scrape_rig_count.py                  # discover both, fetch
    python scripts/scrape_rig_count.py --no-history     # 2024 onwards only
    python scripts/scrape_rig_count.py --url <xlsx> --history <xlsx>
    python scripts/scrape_rig_count.py --dry-run
"""

from __future__ import annotations

import argparse
import csv
import io
import logging
import re
import sys
import urllib.request
from pathlib import Path

logger = logging.getLogger("scrape_rig_count")

INDEX_URL = "https://rigcount.bakerhughes.com/intl-rig-count"
COUNTRY = "NIGERIA"
OUTPUT = "rig_count_nigeria.csv"
COLUMNS = ["month", "land_rigs", "inland_water_rigs", "offshore_rigs", "total_rigs"]

# Baker Hughes drops a plain client outright -- three requests with a short
# User-Agent timed out or had the connection reset before anything came back.
# A full browser header set is answered normally.
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

_STATIC_RE = re.compile(r"/static-files/[a-f0-9-]{36}")
_FILENAME_RE = re.compile(r'filename="([^"]+)"')
# "July-2026  WorldWide Rig Count Report.xlsx" -- note the double space, which
# is why this matches loosely rather than on an exact string.
_WORKBOOK_RE = re.compile(r"worldwide\s+rig\s+count\s+report", re.I)
# The current workbook only reaches back to January 2024. Everything before
# that is in a separate historical file, in a different sheet and a different
# unit -- see parse_master_data.
_HISTORY_RE = re.compile(r"international\s+rig\s+counts\s+for", re.I)
_MONTH_IN_NAME_RE = re.compile(r"([A-Za-z]+)-(\d{4})")

_MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}


def _get(url: str, timeout: int = 60) -> tuple[bytes, dict]:
    request = urllib.request.Request(url, headers=_HEADERS)
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read(), dict(response.headers)


def discover(pattern: "re.Pattern", label: str) -> tuple[str, str] | None:
    """The newest matching workbook's URL and filename, read off the page.

    Every /static-files link is checked rather than the first: the same page
    also serves a methodology PDF, an overview PDF, and two historical
    workbooks, and their order is not guaranteed.
    """
    page, _ = _get(INDEX_URL)
    candidates = sorted(set(_STATIC_RE.findall(page.decode("utf-8", "ignore"))))
    if not candidates:
        raise SystemExit(
            f"no /static-files links found on {INDEX_URL}. The page layout has "
            "changed, or the request was blocked."
        )

    found: list[tuple[tuple[int, int], str, str]] = []
    for path in candidates:
        url = f"https://rigcount.bakerhughes.com{path}"
        try:
            request = urllib.request.Request(url, headers=_HEADERS, method="HEAD")
            with urllib.request.urlopen(request, timeout=30) as response:
                disposition = response.headers.get("Content-Disposition", "")
        except Exception as exc:
            logger.warning("could not check %s: %s", path, exc)
            continue
        match = _FILENAME_RE.search(disposition)
        if not match or not pattern.search(match.group(1)):
            continue
        name = match.group(1)
        # Sorted by the month IN THE FILENAME, so "July-2026" beats
        # "July-2025" no matter what order the page lists them in.
        when = _MONTH_IN_NAME_RE.search(name)
        key = (
            (int(when.group(2)), _MONTHS.get(when.group(1).lower(), 0))
            if when
            else (0, 0)
        )
        found.append((key, url, name))
        logger.info("candidate (%s): %s", label, name)

    if not found:
        logger.warning("no %s workbook among %d files on %s", label, len(candidates), INDEX_URL)
        return None
    _, url, name = max(found)
    logger.info("using %s", name)
    return url, name


def _nigeria_rows(sheet) -> "list[tuple]":
    """Nigeria's rows from a long-format sheet, header found rather than assumed.

    Both workbooks put a summary block above the real table, so rows are
    skipped until the header appears. The two sheets name their fifth column
    differently -- "Rig Status" against "WellStatus" -- but the first eight
    columns are otherwise in the same order, which is why only the first two
    are matched on.
    """
    rows = []
    seen_header = False
    for row in sheet.iter_rows(values_only=True):
        if not seen_header:
            seen_header = bool(row and row[0] == "Region" and row[1] == "Country")
            continue
        if row and row[1] and str(row[1]).upper() == COUNTRY:
            rows.append(row)
    return rows


def _accumulate(rows, per_month: dict, divide_by_fridays: bool) -> None:
    for row in rows:
        _, _, _, location, _, year, month, count = row[:8]
        if not year or not month:
            continue
        weeks = 1
        if divide_by_fridays:
            # The historical sheet stores each month's WEEKLY counts summed,
            # not the monthly average: January 2024 land reads 37 there and 9
            # in the current workbook, and 37/4 Fridays is 9.25. Dividing
            # reproduces the published figure for every overlapping month.
            weeks = int(row[9] or 0) if len(row) > 9 else 0
            if not weeks:
                continue
        key = (int(year), int(month))
        # Lower-cased because the same workbook writes both "Inland Waters"
        # and "Inland waters".
        where = str(location).strip().lower()
        bucket = per_month.setdefault(key, {})
        # Summed: a month carries a row per DrillFor and rig status.
        bucket[where] = bucket.get(where, 0) + (count or 0) / weeks


def parse_workbook(data: bytes, history: bytes | None = None) -> list[dict]:
    """The Nigeria series, newest workbook over historical where they overlap.

    Inland waters is counted, not dropped. Nigeria ran one inland rig every
    month from January to November 2024 and the earlier years carry more, so
    treating the total as land plus offshore quietly understates it.
    """
    import openpyxl

    per_month: dict[tuple[int, int], dict[str, float]] = {}

    if history is not None:
        book = openpyxl.load_workbook(io.BytesIO(history), data_only=True, read_only=True)
        if "Master Data" not in book.sheetnames:
            raise SystemExit(
                f"no 'Master Data' sheet in the historical workbook "
                f"(found {book.sheetnames})."
            )
        _accumulate(_nigeria_rows(book["Master Data"]), per_month, divide_by_fridays=True)
        logger.info("history: %d months", len(per_month))

    book = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if "WW Monthly" not in book.sheetnames:
        raise SystemExit(
            f"no 'WW Monthly' sheet in the workbook (found {book.sheetnames}). "
            "Baker Hughes has changed its layout."
        )
    # Applied second and allowed to overwrite: where the two overlap the
    # current workbook is the published figure and the historical one is
    # derived, so the published figure wins.
    current: dict[tuple[int, int], dict[str, float]] = {}
    _accumulate(_nigeria_rows(book["WW Monthly"]), current, divide_by_fridays=False)
    per_month.update(current)

    if not per_month:
        raise SystemExit(f"no rows for {COUNTRY} in the workbook.")

    rows = []
    for (year, month), values in sorted(per_month.items()):
        land = values.get("land", 0)
        inland = values.get("inland waters", 0)
        offshore = values.get("offshore", 0)
        # Rounded to whole rigs to match how the current workbook publishes
        # them; only the derived historical months are ever fractional.
        rows.append(
            {
                "month": f"{year}-{month:02d}",
                "land_rigs": round(land),
                "inland_water_rigs": round(inland),
                "offshore_rigs": round(offshore),
                "total_rigs": round(land) + round(inland) + round(offshore),
            }
        )
    return rows


def check_sane(rows: list[dict]) -> None:
    """Stop on figures that cannot be a Nigerian rig count.

    Deliberately wide: this is a guard against reading the wrong column or the
    wrong country, not a forecast. Nigeria has run between 9 and 18 rigs over
    the period this file covers, so anything past 200 means the parse is
    wrong, not that drilling boomed.
    """
    bad = [r for r in rows if not 0 <= r["total_rigs"] <= 200]
    if bad:
        raise SystemExit(
            f"{len(bad)} month(s) with an implausible rig count, e.g. "
            f"{bad[0]['month']}={bad[0]['total_rigs']}. Refusing to publish: "
            "this usually means the sheet layout moved and the wrong column "
            "was read."
        )


def check_not_shrinking(rows: list[dict], path: Path, allow: bool) -> None:
    """Refuse to drop months the published file already has.

    The workbook is rewritten whole on every run, so an older one would
    quietly truncate the series while looking like a clean update.
    """
    if not path.exists():
        return
    with path.open(newline="", encoding="utf-8") as handle:
        existing = {r["month"] for r in csv.DictReader(handle) if r.get("month")}
    missing = sorted(existing - {r["month"] for r in rows})
    if not missing:
        return
    logger.warning("%s has months this workbook does not: %s", path.name,
                   ", ".join(missing))
    if not allow:
        raise SystemExit(
            f"{path.name} would lose {len(missing)} month(s): {', '.join(missing)}.\n"
            "NOTHING WAS WRITTEN. This usually means an older workbook was "
            "passed. Pass --allow-fewer-months only if you mean it."
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", help="workbook URL or local path; default: discover")
    parser.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent.parent),
        help="directory to write the CSV into",
    )
    parser.add_argument(
        "--history",
        help="historical workbook (International Rig Counts ...); default: discover",
    )
    parser.add_argument(
        "--no-history",
        action="store_true",
        help="current workbook only, 2024 onwards",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--allow-fewer-months", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    def load(where: str | None, pattern, label):
        """A workbook from a local path, a URL, or discovery."""
        if where and Path(where).exists():
            return Path(where).read_bytes(), Path(where).name
        if where:
            return _get(where)[0], Path(where).name
        hit = discover(pattern, label)
        if hit is None:
            return None, None
        return _get(hit[0])[0], hit[1]

    data, source = load(args.url, _WORKBOOK_RE, "current")
    if data is None:
        raise SystemExit(
            "could not find the current WorldWide Rig Count Report. Pass --url."
        )

    history = history_source = None
    if not args.no_history:
        history, history_source = load(args.history, _HISTORY_RE, "history")
        if history is None:
            logger.warning("no historical workbook; series starts at 2024-01")

    # An xlsx is a zip; anything else means a login page or an error body was
    # served instead, which openpyxl would report far less clearly.
    if not data.startswith(b"PK"):
        raise SystemExit(
            f"{source} is not an xlsx ({len(data)} bytes). Baker Hughes may have "
            "blocked the request."
        )

    if history is not None and not history.startswith(b"PK"):
        raise SystemExit(f"{history_source} is not an xlsx.")
    rows = parse_workbook(data, history)
    check_sane(rows)

    path = Path(args.out) / OUTPUT
    check_not_shrinking(rows, path, args.allow_fewer_months)

    latest = rows[-1]
    logger.info(
        "%d months, %s to %s; latest %s rigs (%s land, %s offshore)",
        len(rows), rows[0]["month"], latest["month"],
        latest["total_rigs"], latest["land_rigs"], latest["offshore_rigs"],
    )

    if args.dry_run:
        logger.info("dry run, nothing written")
        for row in rows[-6:]:
            print(",".join(str(row[c]) for c in COLUMNS))
        return 0

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=COLUMNS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    path.write_text(buffer.getvalue(), encoding="utf-8", newline="")
    logger.info("wrote %s (from %s%s)", path, source,
                f" + {history_source}" if history_source else "")
    return 0


if __name__ == "__main__":
    sys.exit(main())
